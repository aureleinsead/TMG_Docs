import json
import tempfile
import zipfile
import os
import shutil
import time
from pathlib import Path
from typing import Any, Dict

import pandas as pd


# =========================
# LOADERS
# =========================

def load_layout_text(layout_path: str) -> str:
    raw = Path(layout_path).read_bytes()

    for enc in ["utf-16-le", "utf-8-sig", "utf-8", "latin-1"]:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue

    raise ValueError(f"Unable to decode file: {layout_path}")


def load_layout_json(layout_path: str) -> Dict[str, Any]:
    return json.loads(load_layout_text(layout_path))


def extract_layout_from_zip(zip_path: str, tmp_dir: str) -> str:
    """Extract the Report/Layout file from a .zip or .pbix archive."""
    with zipfile.ZipFile(zip_path, "r") as z:
        candidates = [
            n for n in z.namelist()
            if n.lower().endswith("/layout") or n.lower() == "layout"
               or n.lower().endswith("/report/layout")
               or n.lower() == "report/layout"
        ]
        if not candidates:
            raise FileNotFoundError(
                f"No 'Report/Layout' file found inside {zip_path}. "
                f"Contents: {z.namelist()[:20]}"
            )
        z.extract(candidates[0], tmp_dir)
        return str(Path(tmp_dir) / candidates[0])


def safe_json_loads(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except Exception:
        return value


# =========================
# SUPPORTING MAPS
# =========================

def build_alias_map(obj, alias_map=None):
    """Walk query/config JSON to build a map from From-clause alias -> entity name."""
    if alias_map is None:
        alias_map = {}

    if isinstance(obj, dict):
        from_clause = obj.get("From")
        if isinstance(from_clause, list):
            for entry in from_clause:
                if isinstance(entry, dict):
                    alias = entry.get("Name")
                    entity = entry.get("Entity")
                    if alias and entity:
                        alias_map[alias] = entity

        for v in obj.values():
            build_alias_map(v, alias_map)

    elif isinstance(obj, list):
        for item in obj:
            build_alias_map(item, alias_map)

    return alias_map


def build_field_kind_map(obj, kind_map=None):
    """Walk query/config JSON to build a map from Select Name -> (kind, function)."""
    if kind_map is None:
        kind_map = {}

    if isinstance(obj, dict):
        name = obj.get("Name")
        if name and isinstance(name, str):
            if "Aggregation" in obj and isinstance(obj["Aggregation"], dict):
                kind_map[name] = ("aggregation", obj["Aggregation"].get("Function"))
            elif "Measure" in obj and isinstance(obj["Measure"], dict):
                kind_map[name] = ("measure", None)
            elif "HierarchyLevel" in obj and isinstance(obj["HierarchyLevel"], dict):
                kind_map[name] = ("hierarchy_level", None)
            elif "Column" in obj and isinstance(obj["Column"], dict):
                kind_map[name] = ("column", None)

        for v in obj.values():
            build_field_kind_map(v, kind_map)

    elif isinstance(obj, list):
        for item in obj:
            build_field_kind_map(item, kind_map)

    return kind_map


def build_hierarchy_display_map(obj, h_map=None):
    """Build map from Select Name -> composed 'SourceColumn LevelName' for HierarchyLevel fields."""
    if h_map is None:
        h_map = {}
    if isinstance(obj, dict):
        name = obj.get("Name")
        if name and isinstance(name, str) and "HierarchyLevel" in obj:
            hl = obj["HierarchyLevel"]
            if isinstance(hl, dict):
                level = hl.get("Level", "")
                # Try to get the source column name from the queryName pattern:
                # {table}.{column}.Variation.{hierarchy}.{level}
                if ".Variation." in name and level:
                    before_var = name.split(".Variation.")[0]
                    # Column name is after the first dot (table.column)
                    if "." in before_var:
                        col_name = before_var.split(".", 1)[1]
                        h_map[name] = f"{col_name} {level}"
                    else:
                        h_map[name] = f"{before_var} {level}"
                elif level:
                    # Fallback: try Expression hierarchy structure
                    hierarchy_expr = hl.get("Expression", {}).get("Hierarchy", {})
                    if isinstance(hierarchy_expr, dict):
                        h_name = hierarchy_expr.get("Hierarchy", "")
                        if h_name:
                            h_map[name] = f"{h_name} {level}"
        for v in obj.values():
            build_hierarchy_display_map(v, h_map)
    elif isinstance(obj, list):
        for item in obj:
            build_hierarchy_display_map(item, h_map)
    return h_map


# =========================
# ROLE EXTRACTION
# =========================

def extract_roles_from_query(query_json, role_map=None):
    if role_map is None:
        role_map = {}

    try:
        if isinstance(query_json, dict):
            selects = query_json.get("Select")
            if isinstance(selects, list):
                for item in selects:
                    name = item.get("Name")
                    roles = item.get("Roles", {})
                    for role_name, is_active in roles.items():
                        if is_active and name:
                            role_map.setdefault(name, []).append(role_name)

            for v in query_json.values():
                extract_roles_from_query(v, role_map)

        elif isinstance(query_json, list):
            for item in query_json:
                extract_roles_from_query(item, role_map)

    except Exception:
        pass

    return role_map


def extract_roles_from_projections(config_json):
    role_map = {}

    try:
        projections = config_json.get("singleVisual", {}).get("projections", {})

        for role, items in projections.items():
            if not isinstance(items, list):
                continue

            for item in items:
                ref = item.get("queryRef")
                if ref:
                    role_map.setdefault(ref, []).append(role)

    except Exception:
        pass

    return role_map


def extract_column_properties(config_json):
    """Extract visual-level column renames from singleVisual.columnProperties."""
    rename_map = {}
    try:
        col_props = config_json.get("singleVisual", {}).get("columnProperties", {})
        if isinstance(col_props, dict):
            for qref, props in col_props.items():
                if isinstance(props, dict):
                    dn = props.get("displayName")
                    if dn:
                        rename_map[qref] = dn
    except Exception:
        pass
    return rename_map


# =========================
# FIELD RESOLUTION HELPERS
# =========================

AGG_PREFIXES = ["Sum(", "Avg(", "Count(", "CountNonNull(", "Min(", "Max(",
                "StDev(", "Var(", "Median("]


def _strip_agg_wrapper(ref):
    """Strip aggregation function wrapper: 'Sum(m.Amount)' -> 'm.Amount'."""
    for prefix in AGG_PREFIXES:
        if ref.startswith(prefix) and ref.endswith(")"):
            return ref[len(prefix):-1]
    return ref


def _resolve_technical_name(ref, alias_map):
    """Convert a queryRef to 'EntityName.Property' using the alias map, preserving any aggregation wrapper."""
    inner = _strip_agg_wrapper(ref)
    if "." in inner:
        alias, prop = inner.split(".", 1)
        entity = alias_map.get(alias, alias)
        resolved = f"{entity}.{prop}"
        # Preserve aggregation wrapper if present
        if inner != ref:
            wrapper_prefix = ref[:len(ref) - len(inner) - 1]
            return f"{wrapper_prefix}{resolved})"
        return resolved
    return ref


def _display_from_ref(ref):
    """Extract display name from a queryRef string."""
    inner = _strip_agg_wrapper(ref)
    if "." in inner:
        return inner.split(".", 1)[1]
    return inner


def _resolve_kind(ref, kind_map):
    """Determine field kind from kind_map."""
    if ref in kind_map:
        kind_type, func = kind_map[ref]
        if kind_type == "measure":
            return "Measure"
        if kind_type == "hierarchy_level":
            return "HierarchyLevel"
        if kind_type == "aggregation":
            return f"Aggregation({func})"
    return "Column"


# =========================
# ROLE FORMAT
# =========================

def format_roles(roles):
    ROLE_ORDER = ["Values", "Rows", "Columns", "Category", "Series", "Tooltips"]

    unique = list(dict.fromkeys([r for r in roles if r]))

    return " | ".join(
        sorted(unique, key=lambda x: ROLE_ORDER.index(x) if x in ROLE_ORDER else 99)
    )


# =========================
# VISUAL HELPERS
# =========================

def unique_join(values):
    seen = set()
    out = []
    for v in values:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return ", ".join(out)


def get_visual_type(config_json):
    return config_json.get("singleVisual", {}).get("visualType")


def get_visual_title(config_json):
    try:
        sv = config_json.get("singleVisual", {})

        vc_objects = sv.get("vcObjects")
        if isinstance(vc_objects, dict):
            title = vc_objects.get("title", [{}])[0]
            val = (
                title.get("properties", {})
                     .get("text", {})
                     .get("expr", {})
                     .get("Literal", {})
                     .get("Value")
            )
            if isinstance(val, str) and val.strip():
                return val.strip("'\"")

        objects = sv.get("objects")
        if isinstance(objects, dict):
            title = objects.get("title", [{}])[0]
            val = (
                title.get("properties", {})
                     .get("text", {})
                     .get("expr", {})
                     .get("Literal", {})
                     .get("Value")
            )
            if isinstance(val, str) and val.strip():
                return val.strip("'\"")

        return None

    except Exception:
        return None


def extract_textbox_content(config_json):
    try:
        sv = config_json.get("singleVisual", {})

        paragraphs = (
            sv.get("objects", {})
              .get("general", [{}])[0]
              .get("properties", {})
              .get("paragraphs", [])
        )

        # Use only the first non-empty paragraph
        for p in paragraphs:
            parts = []
            for run in p.get("textRuns", []):
                val = run.get("value")
                if isinstance(val, str) and val.strip():
                    parts.append(val.strip())
            line = " ".join(parts)
            if line:
                return line

        return None

    except Exception:
        return None


# =========================
# VISUAL CORE
# =========================

def extract_visual_container_info(page_display_name, visual_container, visual_index):

    config = safe_json_loads(visual_container.get("config"))
    query = safe_json_loads(visual_container.get("query"))
    data_transforms = safe_json_loads(visual_container.get("dataTransforms"))

    # --- Visual metadata ---
    visual_type = None
    visual_title = None

    if isinstance(config, dict):
        visual_type = get_visual_type(config)
        if visual_type == "textbox":
            base_title = extract_textbox_content(config)
        else:
            base_title = get_visual_title(config)

        fallback_name = visual_type if isinstance(visual_type, str) and visual_type.strip() else ""
        visual_title = base_title or (f"{fallback_name} [{visual_index}]" if fallback_name else f"[{visual_index}]")

    # --- Build supporting maps ---
    alias_map = {}
    kind_map = {}
    hierarchy_map = {}
    for source in [query, config]:
        if isinstance(source, dict):
            build_alias_map(source, alias_map)
            build_field_kind_map(source, kind_map)
            build_hierarchy_display_map(source, hierarchy_map)

    # Role map: projections (highest priority) overriding query roles
    # Each entry is a list of roles
    role_map = {}
    if isinstance(query, dict):
        extract_roles_from_query(query, role_map)
    if isinstance(config, dict):
        proj_roles = extract_roles_from_projections(config)
        for k, v in proj_roles.items():
            role_map.setdefault(k, []).extend(v)

    # Column properties: visual-level display name overrides
    col_rename_map = {}
    if isinstance(config, dict):
        col_rename_map = extract_column_properties(config)

    # --- Build field list (no duplicates) ---
    enriched = []
    roles = []
    seen = set()
    seen_display = set()  # Deduplicate by display name within visual

    # Primary source: dataTransforms.selects (one entry per field, with displayName)
    dt_selects = []
    if isinstance(data_transforms, dict):
        sel_list = data_transforms.get("selects")
        if isinstance(sel_list, list):
            dt_selects = sel_list

    if dt_selects:
        # Two-pass approach: collect all candidates, then select best per field
        candidates = []
        seen_qnames = set()
        for sel in dt_selects:
            if not isinstance(sel, dict):
                continue
            qname = sel.get("queryName", "")
            if not qname or qname in seen_qnames:
                continue
            seen_qnames.add(qname)

            # Display name: columnProperties > hierarchy composition > dataTransforms > parsed
            display = col_rename_map.get(qname) or sel.get("displayName") or _display_from_ref(qname)

            technical = _resolve_technical_name(qname, alias_map)
            kind = _resolve_kind(qname, kind_map)

            # For HierarchyLevel, compose display from hierarchy + level
            if kind == "HierarchyLevel" and qname in hierarchy_map:
                display = col_rename_map.get(qname) or hierarchy_map[qname]

            # For Measures, use just the property name (no table prefix)
            if kind == "Measure":
                inner = _strip_agg_wrapper(technical)
                if "." in inner:
                    technical = inner.split(".", 1)[1]

            # Role
            role_list = role_map.get(qname, [])
            if not role_list:
                dt_roles = sel.get("roles")
                if isinstance(dt_roles, dict):
                    role_list = [rname for rname, active in dt_roles.items() if active]
            role = ", ".join(dict.fromkeys(role_list)) if role_list else ""

            is_wrapped = _strip_agg_wrapper(technical) != technical
            stripped_tech = _strip_agg_wrapper(technical)

            candidates.append({
                "display": display,
                "technical": technical,
                "kind": kind,
                "role": role,
                "is_wrapped": is_wrapped,
                "stripped_tech": stripped_tech,
            })

        # Group by stripped technical name, select best entry per group
        from collections import OrderedDict
        groups = OrderedDict()
        for entry in candidates:
            groups.setdefault(entry["stripped_tech"], []).append(entry)

        for stripped, group in groups.items():
            # Prefer non-wrapped entries; among those, keep the last one
            non_wrapped = [e for e in group if not e["is_wrapped"]]
            if non_wrapped:
                selected = non_wrapped[-1]
            else:
                selected = group[-1]

            display = selected["display"]
            if display in seen_display:
                continue
            seen_display.add(display)

            role = selected["role"]
            if role:
                roles.append(role)

            enriched.append({
                "display": display,
                "technical": selected["technical"],
                "kind": selected["kind"],
                "role": role
            })

    elif role_map:
        # Fallback: use projections directly (no dataTransforms available)
        for ref, role_list in role_map.items():
            if ref in seen:
                continue
            seen.add(ref)

            display = col_rename_map.get(ref) or _display_from_ref(ref)
            technical = _resolve_technical_name(ref, alias_map)
            kind = _resolve_kind(ref, kind_map)

            # For Measures, use just the property name (no table prefix)
            if kind == "Measure":
                inner = _strip_agg_wrapper(technical)
                if "." in inner:
                    technical = inner.split(".", 1)[1]

            role = ", ".join(dict.fromkeys(role_list)) if role_list else ""
            if role:
                roles.append(role)

            enriched.append({
                "display": display,
                "technical": technical,
                "kind": kind,
                "role": role
            })

    # --- Build output rows ---
    visual_row = {
        "Page": page_display_name,
        "Visual Index": visual_index,
        "Visual Name": visual_title,
        "Visual Type": visual_type,
        "Field Count": len(enriched),
        "Fields (display)": unique_join([e["display"] for e in enriched if e["display"]]),
        "Fields (technical)": unique_join([e["technical"] for e in enriched if e["technical"]]),
        "Projection Roles": format_roles(roles),
        "Position X": visual_container.get("x"),
        "Position Y": visual_container.get("y"),
        "Width": visual_container.get("width"),
        "Height": visual_container.get("height"),
    }

    field_output = [
        {
            "Page": page_display_name,
            "Visual Index": visual_index,
            "Visual Name": visual_title,
            "Visual Type": visual_type,
            "Field Display Name": e["display"],
            "Field Technical Name": e["technical"],
            "Field Kind": e["kind"],
            "Projection Role": e["role"],
        }
        for e in enriched
    ]

    return visual_row, field_output


# =========================
# MAIN + EXPORT (UNCHANGED)
# =========================

def extract_layout_inventory(layout_path):

    layout = load_layout_json(layout_path)
    sections = layout.get("sections", [])

    page_rows = []
    visual_rows = []
    field_rows = []

    for section in sections:

        page_display_name = section.get("displayName") or section.get("name")
        vcs = section.get("visualContainers", [])

        page_visuals = []
        page_fields = []

        for i, vc in enumerate(vcs, start=1):
            v, f = extract_visual_container_info(page_display_name, vc, i)
            visual_rows.append(v)
            field_rows.extend(f)

            page_visuals.append(v)
            page_fields.extend(f)

        page_rows.append({
            "Page": page_display_name,
            "Visuals": len(page_visuals),
            "Data Visuals": sum(1 for v in page_visuals if v["Field Count"] > 0),
            "Total Fields": len(page_fields),
        })

    df_pages = pd.DataFrame(page_rows)
    df_visuals = pd.DataFrame(visual_rows)
    df_fields = pd.DataFrame(field_rows)

    df_data_visuals_only = df_visuals[df_visuals["Field Count"] > 0].copy()

    return df_pages, df_visuals, df_fields, df_data_visuals_only


# =========================
# PBIR FORMAT (no Layout file)
# =========================

def _extract_fields_from_pbir_query_state(query_state):
    """Build field list from PBIR queryState: {Role: {projections: [...]}}."""
    if not isinstance(query_state, dict):
        return []

    fields = []
    seen = set()

    for role_name, role_data in query_state.items():
        if not isinstance(role_data, dict):
            continue
        for proj in role_data.get("projections", []):
            if not isinstance(proj, dict):
                continue
            if not proj.get("active", True):
                continue

            query_ref = proj.get("queryRef", "")
            field = proj.get("field", {})
            display = (proj.get("displayName")
                       or proj.get("nativeQueryRef")
                       or (query_ref.split(".", 1)[1] if "." in query_ref else query_ref))

            if "Column" in field:
                kind = "Column"
                entity = field["Column"].get("Expression", {}).get("SourceRef", {}).get("Entity", "")
                prop = field["Column"].get("Property", "")
                technical = f"{entity}.{prop}" if entity else prop
            elif "Measure" in field:
                kind = "Measure"
                technical = field["Measure"].get("Property", query_ref)
            elif "HierarchyLevel" in field:
                kind = "HierarchyLevel"
                technical = query_ref
            else:
                kind = "Column"
                technical = query_ref

            key = (technical, role_name)
            if key in seen:
                continue
            seen.add(key)

            fields.append({"display": display, "technical": technical,
                           "kind": kind, "role": role_name})

    return fields


def _extract_visual_container_info_pbir(page_display_name, visual_json, visual_index):
    """Extract visual info directly from a PBIR visual.json dict."""
    position = visual_json.get("position", {})
    visual = visual_json.get("visual", {})
    visual_type = visual.get("visualType")
    query = visual.get("query", {})
    query_state = query.get("queryState", {}) if isinstance(query, dict) else {}

    # Reuse existing title helpers by wrapping visual in singleVisual
    config_wrap = {"singleVisual": visual}
    if visual_type == "textbox":
        visual_title = extract_textbox_content(config_wrap)
    else:
        visual_title = get_visual_title(config_wrap)
    fallback = visual_type or ""
    visual_title = visual_title or (f"{fallback} [{visual_index}]" if fallback else f"[{visual_index}]")

    fields = _extract_fields_from_pbir_query_state(query_state)
    roles = [f["role"] for f in fields if f["role"]]

    visual_row = {
        "Page": page_display_name,
        "Visual Index": visual_index,
        "Visual Name": visual_title,
        "Visual Type": visual_type,
        "Field Count": len(fields),
        "Fields (display)": unique_join([f["display"] for f in fields if f["display"]]),
        "Fields (technical)": unique_join([f["technical"] for f in fields if f["technical"]]),
        "Projection Roles": format_roles(roles),
        "Position X": position.get("x"),
        "Position Y": position.get("y"),
        "Width": position.get("width"),
        "Height": position.get("height"),
    }

    field_output = [
        {
            "Page": page_display_name,
            "Visual Index": visual_index,
            "Visual Name": visual_title,
            "Visual Type": visual_type,
            "Field Display Name": f["display"],
            "Field Technical Name": f["technical"],
            "Field Kind": f["kind"],
            "Projection Role": f["role"],
        }
        for f in fields
    ]

    return visual_row, field_output


def extract_pbir_inventory(report_dir: Path):
    """Extract layout inventory from a PBIR-format Report directory."""
    # Locate page.json files regardless of how deep pages/ is nested
    page_json_paths = list(report_dir.rglob("page.json"))
    if not page_json_paths:
        raise FileNotFoundError(f"No PBIR page structure (page.json) found inside {report_dir}")

    page_entries = []
    for page_json_path in page_json_paths:
        try:
            page_data = json.loads(page_json_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        page_folder = page_json_path.parent
        page_entries.append((page_data.get("ordinal", 0), page_folder, page_data))

    page_entries.sort(key=lambda x: x[0])

    page_rows = []
    visual_rows = []
    field_rows = []

    for _, page_folder, page_data in page_entries:
        page_display_name = page_data.get("displayName") or page_data.get("name") or page_folder.name

        visuals_dir = page_folder / "visuals"
        visual_entries = []
        if visuals_dir.is_dir():
            for vis_folder in sorted(visuals_dir.iterdir()):
                if not vis_folder.is_dir():
                    continue
                vis_json_path = vis_folder / "visual.json"
                if not vis_json_path.exists():
                    continue
                try:
                    vis_data = json.loads(vis_json_path.read_text(encoding="utf-8"))
                except Exception:
                    continue
                pos = vis_data.get("position", {})
                visual_entries.append((pos.get("z", 0), pos.get("tabOrder", 0), vis_data))

        visual_entries.sort(key=lambda x: (x[0], x[1]))

        page_visuals = []
        page_fields = []
        for i, (_, _, vis_data) in enumerate(visual_entries, start=1):
            v, f = _extract_visual_container_info_pbir(page_display_name, vis_data, i)
            visual_rows.append(v)
            field_rows.extend(f)
            page_visuals.append(v)
            page_fields.extend(f)

        page_rows.append({
            "Page": page_display_name,
            "Visuals": len(page_visuals),
            "Data Visuals": sum(1 for v in page_visuals if v["Field Count"] > 0),
            "Total Fields": len(page_fields),
        })

    df_pages = pd.DataFrame(page_rows)
    df_visuals = pd.DataFrame(visual_rows)
    df_fields = pd.DataFrame(field_rows)
    df_data_visuals_only = df_visuals[df_visuals["Field Count"] > 0].copy()

    return df_pages, df_visuals, df_fields, df_data_visuals_only


def export_to_excel(layout_path, output_xlsx):

    df_pages, df_visuals, df_fields, df_data_visuals_only = extract_layout_inventory(layout_path)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_pages.to_excel(writer, sheet_name="PAGES_SUMMARY", index=False)
        df_visuals.to_excel(writer, sheet_name="VISUAL_INVENTORY", index=False)
        df_fields.to_excel(writer, sheet_name="FIELDS_BY_VISUAL", index=False)
        df_data_visuals_only.to_excel(writer, sheet_name="DATA_VISUALS_ONLY", index=False)

        workbook = writer.book

        # =========================
        # STYLING FUNCTION
        # =========================
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        def style_sheet(sheet_name):
            ws = workbook[sheet_name]

            # Freeze header row
            ws.freeze_panes = "A2"

            # Enable filter on header row
            ws.auto_filter.ref = ws.dimensions

            # Wrap text + auto width
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        # Apply styling to all sheets
        for sheet in workbook.sheetnames:
            style_sheet(sheet)

    return output_xlsx


def export_pbir_to_excel(report_dir, output_xlsx):

    df_pages, df_visuals, df_fields, df_data_visuals_only = extract_pbir_inventory(report_dir)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_pages.to_excel(writer, sheet_name="PAGES_SUMMARY", index=False)
        df_visuals.to_excel(writer, sheet_name="VISUAL_INVENTORY", index=False)
        df_fields.to_excel(writer, sheet_name="FIELDS_BY_VISUAL", index=False)
        df_data_visuals_only.to_excel(writer, sheet_name="DATA_VISUALS_ONLY", index=False)

        workbook = writer.book

        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment

        def style_sheet(sheet_name):
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        for sheet in workbook.sheetnames:
            style_sheet(sheet)

    return output_xlsx
# =========================
# RUN
# =========================

if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()

    # Ask user to choose between a .zip/.pbix archive or an already-extracted folder
    choice = tk.StringVar(value="")

    def _pick_zip():
        choice.set("zip")
        picker.destroy()

    def _pick_folder():
        choice.set("folder")
        picker.destroy()

    picker = tk.Toplevel(root)
    picker.title("Power BI Extractor - Select input type")
    picker.resizable(False, False)
    tk.Label(picker, text="What do you want to open?", font=("Segoe UI", 10)).pack(padx=24, pady=(16, 6))
    tk.Button(picker, text="Select a .zip / .pbix file", command=_pick_zip, width=30).pack(padx=24, pady=4)
    tk.Button(picker, text="Select an extracted folder",  command=_pick_folder, width=30).pack(padx=24, pady=(4, 16))
    picker.protocol("WM_DELETE_WINDOW", lambda: (choice.set("cancel"), picker.destroy()))
    picker.grab_set()
    root.wait_window(picker)

    if choice.get() in ("", "cancel"):
        print("No input selected.")
        root.destroy()
        exit()

    # If user selected a .zip or .pbix archive, extract and search for Report/Layout
    tmp_dir = None
    try:
        if choice.get() == "zip":
            layout_file = filedialog.askopenfilename(
                title="Select Power BI .zip or .pbix archive",
                filetypes=[("Power BI / ZIP archives", "*.zip *.pbix"), ("All files", "*.*")]
            )

            if not layout_file:
                print("No file selected.")
                exit()

            tmp_dir = tempfile.mkdtemp(prefix="pbi_extract_")
            try:
                with zipfile.ZipFile(layout_file, "r") as z:
                    z.extractall(tmp_dir)
            except Exception as e:
                root.destroy()
                print(f"Error extracting archive: {e}")
                time.sleep(10)
                if tmp_dir and os.path.exists(tmp_dir):
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                exit(1)

            search_root = Path(tmp_dir)

        else:  # folder
            selected_folder = filedialog.askdirectory(
                title="Select extracted Power BI folder"
            )

            if not selected_folder:
                print("No folder selected.")
                exit()

            search_root = Path(selected_folder)

        # Search for a 'Report' directory (case-insensitive)
        report_dirs = [p for p in search_root.rglob("*") if p.is_dir() and p.name.lower() == "report"]
        if not report_dirs:
            root.destroy()
            print(f"Error: 'Report' folder not found inside {search_root}.")
            time.sleep(10)
            if tmp_dir and os.path.exists(tmp_dir):
                shutil.rmtree(tmp_dir, ignore_errors=True)
            exit(1)

        report_dir = report_dirs[0]

        # Try traditional Layout file first, then fall back to PBIR format
        layout_candidates = [p for p in report_dir.rglob("*") if p.is_file() and p.name.lower() == "layout"]
        if not layout_candidates:
            layout_candidates = [p for p in search_root.rglob("*") if p.is_file() and p.name.lower() == "layout"]

        output_file = filedialog.asksaveasfilename(
            title="Save Excel output file",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")]
        )
        if not output_file:
            output_file = "powerbi_visual_field_extraction.xlsx"

        if layout_candidates:
            # Traditional Layout format
            try:
                result = export_to_excel(str(layout_candidates[0]), output_file)
                print(f"Extraction complete: {result}")
            except Exception as e:
                print(f"Error: {e}")
                root.destroy()
                time.sleep(10)
                if tmp_dir and os.path.exists(tmp_dir):
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                exit(1)

        elif list(report_dir.rglob("page.json")):
            # PBIR format detected via page.json files
            print("Traditional Layout not found — using PBIR page-based extraction.")
            try:
                result = export_pbir_to_excel(report_dir, output_file)
                print(f"Extraction complete: {result}")
            except Exception as e:
                print(f"Error: {e}")
                root.destroy()
                time.sleep(10)
                if tmp_dir and os.path.exists(tmp_dir):
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                exit(1)

        else:
            root.destroy()
            print(f"Error: No compatible Power BI format found in {search_root}.")
            print("Expected a 'Layout' file (traditional .pbix) or page.json files (PBIR format).")
            print(f"Contents of Report folder: {[p.name for p in report_dir.rglob('*')][:50]}")
            time.sleep(10)
            if tmp_dir and os.path.exists(tmp_dir):
                shutil.rmtree(tmp_dir, ignore_errors=True)
            exit(1)

    finally:
        try:
            root.destroy()
        except Exception:
            pass
        if tmp_dir and os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir, ignore_errors=True)